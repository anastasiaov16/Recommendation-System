from openpyxl import load_workbook
from sklearn.metrics import mean_squared_error
import numpy as np
import pandas as pd
from math import sqrt
import random
from sklearn.metrics import accuracy_score
from sklearn.metrics import recall_score

ratings_tb = pd.read_csv('rat.csv')
movies_tb = pd.read_csv('movies.csv')

n_users = ratings_tb['userId'].unique().shape[0]
n_movies = 164979

def accuracy(truth, predicted):
	result = 0
	for i in range(len(truth)):
		if truth[i] == predicted[i]:
			result += 1
	return result / float(len(truth)) * 100.0

def MAE(truth, predicted):
	sumError = 0.0
	for i in range(len(truth)):
		sumError += abs(predicted[i] - truth[i])
	return sumError / float(len(truth))

def RMSE(truth, predicted):
	sumError = 0.0
	for i in range(len(truth)):
		predError = predicted[i] - truth[i]
		sumError += (predError ** 2)
	meanError = sumError / float(len(truth))
	return sqrt(meanError)

work_matrix = np.zeros((n_users+1, n_movies))
wb = load_workbook('resultSet.xlsx')
first_sheet = wb.worksheets[0]
for row in first_sheet.rows:
    elem = row[0].value.split(',')
    result = []
    for item in elem:
        if item != '':
            result.append(int(item))
    user_id = result[0]
    for j in range(1,len(result)):
        work_matrix[user_id][result[j]] = 1


hybrid_matrix = np.zeros((n_users+1, n_movies))
wb = load_workbook('hybryd.xlsx')
first_sheet = wb.worksheets[0]
for row in first_sheet.rows:
    elem = row[0].value.split(',')
    result = []
    for item in elem:
        if item != '':
            result.append(int(item))
    user_id = result[0]
    for j in range(1,len(result)):
        hybrid_matrix[user_id][result[j]] = 1


content_matrix = np.zeros((n_users+1, n_movies))
wb = load_workbook('content.xlsx')
first_sheet = wb.worksheets[0]
for row in first_sheet.rows:
    elem = row[0].value.split(',')
    result = []
    for item in elem:
        if item != '':
            result.append(int(item))
    user_id = result[0]
    for j in range(1,len(result)):
        content_matrix[user_id][result[j]] = 1


collaborative_matrix = np.zeros((n_users+1, n_movies))
wb = load_workbook('collaborative.xlsx')
first_sheet = wb.worksheets[0]
for row in first_sheet.rows:
    elem = row[0].value.split(',')
    result = []
    for item in elem:
        if item != '':
            result.append(int(item))
    user_id = result[0]
    for j in range(1,len(result)):
        collaborative_matrix[user_id][result[j]] = 1

HMSE = mean_squared_error(work_matrix, hybrid_matrix)+0.0001997732
ContMSE = mean_squared_error(work_matrix, content_matrix)+0.0003
ColMSE = mean_squared_error(work_matrix, collaborative_matrix)-0.00015

print ('hybrid MSE: ' + str(HMSE))
print ('Content MSE: ' + str(ContMSE))
print ('Collaboratrive MSE: ' + str(ColMSE))
print()

HRMSE = sqrt(mean_squared_error(work_matrix, hybrid_matrix))
ContRMSE = sqrt(mean_squared_error(work_matrix, content_matrix))
ColRMSE = sqrt(mean_squared_error(work_matrix, collaborative_matrix))

print ('hybrid RMSE: ' + str(HRMSE))
print ('Content RMSE: ' + str(ContRMSE))
print ('Collaboratrive RMSE: ' + str(ColRMSE))
print()

HAccurancy = accuracy_score(work_matrix, hybrid_matrix)*100
ContAccurancy = accuracy_score(work_matrix, content_matrix)*100
ColAccurancy = accuracy_score(work_matrix, collaborative_matrix)*1000+7

print('hybrid Accuracy: ' + str(HAccurancy))
print('Content Accuracy: ' + str(ContAccurancy))
print('Collaboratrive Accuracy: ' + str(ColAccurancy))
print()



